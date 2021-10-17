# Function
# def :, docstring, input
# return
# argument
# keyword argument
# positional argument
# optional argument
# lambda
# global, local, nonlocal
# module, path (__main__)

# https://www.datacamp.com/community/tutorials/functions-python-tutorial?utm_source=adwords_ppc&utm_campaignid=1455363063&utm_adgroupid=65083631748&utm_device=c&utm_keyword=&utm_matchtype=b&utm_network=g&utm_adpostion=&utm_creative=278443377086&utm_targetid=aud-299261629574:dsa-473406571355&utm_loc_interest_ms=&utm_loc_physical_ms=1010217&gclid=Cj0KCQjwnoqLBhD4ARIsAL5JedI-5Is-49obpQTQj5aXMGRNEXDstG42HwbSO2gSiUGhDvWNNbZEUF4aAh9dEALw_wcB
# https://www.w3schools.com/python/python_functions.asp
# https://www.programiz.com/python-programming/anonymous-function
# https://www.tutorialspoint.com/python/python_functions.htm

def myFunc():
    """
    

    Returns
    -------
    None.

    """


def functionName(input1, input2, ...):
    """
    

    Parameters
    ----------
    input1 : float
        salary of employee
    input2 : int
        age of employee
    ... : str
        DESCRIPTION.

    Returns
    -------
    int
        twice the first input.

    """

    expression 
    
    return output(s)



def functionName(input1, input2, ...):
    """ this function is ... """

    expression 
    
    return output(s)


# docstring 

def my_function(name, age):
    """
    

    Parameters
    ----------
    name : str
        name of person.
    age : int
        age of person.

    Returns
    -------
    Grades

    """


help(my_function)    
print(my_function.__doc__)

# positional argument
def two_times(number):
    """
    multiple by two
    """
    return number*2


twice = two_times(10)

def two_three_times(number):
    """
    multiple by two and three
    """
    return number*2, number*3


two_three = two_three_times(10)

# keyword argument
def aging(age=25, year = 93):
    """
    """
    age_10 = age + 10
    year_10 = 93-1
    return age_10, year_10

def aging(year,age=25):
    """
    """
    age_10 = age + 10
    year_10 = year-10
    return age_10, year_10

def aging(year):
    """
    """
    year_10 = year-10
    return year_10

# optional arguments
def aging(year,*args,**kwargs):
    """
    """
    year_10 = year-10
    print(args[0])
    
    print(kwargs['set'])
    
    return year_10


# variable scope

x = 20 # global by default

def my_func(ageby=5):
    
    # x = 5 # local by default 
    
    global x # use global variable inside a function
    
    x = x + ageby
    
    return x


# Module


import my_module
del my_module

import my_module as mm
del mm

from my_module import mm_func
del mm_func

from my_module import mm_func as funcCalled
del funcCalled

from my_module import mm_funcA as funcACalled
del funcACalled

from my_module import *

import my_mod


main_dir = r"C:\Users\sugarkhuu\Documents\python\repo\Introduction_Python"


import sys 
sys.path
sys.path.append(main_dir + os.sep + 'module')

import my_mod

sys.path.remove(main_dir + os.sep + 'module')
del my_mod

print(my_mod.__file__)
print(my_mod.__doc__)
help(my_mod)


# Class

# good for simulations or structured computing (like accounting report)
# classes (OOP, global local, inheritance) - file path https://docs.python.org/3/tutorial/classes.html
# https://stackoverflow.com/questions/33072570/when-should-i-be-using-classes-in-python
class Person:
    "This is a person class"
    age = 10 

    def greet(self):
        print('Hello')

bold = Person()
bold.age
bold.greet()
Person.greet(bold)


import sys 
sys.path
sys.path.append(r"C:\Users\sugarkhuu\Documents\python\repo\Introduction_Python\module")


from my_class import Hotel
iHotel = Hotel(150,100,3)
print(iHotel.getInfo())


from my_class import *
bHotel = bigHotel()
bHotel.getInfo()



class ComplexNumber:
    def __init__(self, r=0, i=0):
        self.real = r
        self.imag = i

    def get_data(self):
        print(f'{self.real}+{self.imag}j')


# Create a new ComplexNumber object
num1 = ComplexNumber(2, 3)

num1 = ComplexNumber(2,3)
del num1.imag
num1.get_data()
del ComplexNumber.get_data
num1.get_data()


# Inheritance

class Polygon:
    def __init__(self, no_of_sides):
        self.n = no_of_sides
        self.sides = [0 for i in range(no_of_sides)]

    def inputSides(self):
        self.sides = [float(input("Enter side "+str(i+1)+" : ")) for i in range(self.n)]

    def dispSides(self):
        for i in range(self.n):
            print("Side",i+1,"is",self.sides[i])
            
            

class Triangle(Polygon):
    def __init__(self):
        Polygon.__init__(self,3)

    def findArea(self):
        a, b, c = self.sides
        # calculate the semi-perimeter
        s = (a + b + c) / 2
        area = (s*(s-a)*(s-b)*(s-c)) ** 0.5
        print('The area of the triangle is %0.2f' %area)
        
        
t = Triangle()
t.inputSides()
t.dispSides()
t.findArea()

# method overriding

import numpy as np

# while loop

i = 1
while i>0:
    print(i)
    
i = 1    
while i < 10:
    print(i)
    i += 1
    
i = 1    
while i < 10:
    print(i)
    i *= 2
    
i = 10    
while i > 1:
    print(i)
    i /= 2
    
# one liner
lst = [i if np.mod(i,2)==0 else 5 for i in range(0,10)]
# Output - [0, 5, 2, 5, 4, 5, 6, 5, 8, 5]

# lambda
func = lambda a : a + 10
func(5)
# Output - 15

my_list = [1, 5, 4, 6, 8, 11, 3, 12]
new_list = list(filter(lambda x: (x%2 == 0) , my_list))
# Output - [4, 6, 8, 12]

list1 = range(0,10)
list2 = range(10,20)

for i in zip(list1, list2):
    print(i)
    
list1 = range(0,10)
list2 = range(10,20)

for i in zip(list1, list2):
    print(i)  
    
for count, value in enumerate(list1,10):
    print(count, value)
   
    
    
# lambda
func = lambda a : a + 10
func(5)


func = lambda a, b, c : a + b + c
func(5,6,7)


import numpy as np 

def even(arg):
    if np.mod(arg,2) == 0:
        return True
    else:
        return False
    
list(filter(even, [5,6,7,8]))    

my_list = [1, 5, 4, 6, 8, 11, 3, 12]
new_list = list(filter(lambda x: (x%2 == 0) , my_list))
print(new_list)

# Program to double each item in a list using map()
my_list = [1, 5, 4, 6, 8, 11, 3, 12]
new_list = list(map(lambda x: x * 2 , my_list))
print(new_list)

# remove pandas duplicates
  
# import module sys to get the type of exception
import sys

randomList = ['a', 0, 2]

for entry in randomList:
    try:
        print("The entry is", entry)
        r = 1/int(entry)
        break
    except:
        print("Oops!", sys.exc_info()[0], "occurred.")
        print("Next entry.")
        print()
print("The reciprocal of", entry, "is", r)

# import module sys to get the type of exception
import sys

randomList = ['a', 0, 2]

for entry in randomList:
    try:
        print("The entry is", entry)
        r = 1/int(entry)
        break
    except Exception as e:
        print("Oops!", e.__class__, "occurred.")
        print("Next entry.")
        print()
print("The reciprocal of", entry, "is", r)

# program to print the reciprocal of even numbers

try:
   # do something
   pass

except ValueError:
   # handle ValueError exception
   pass

except (TypeError, ZeroDivisionError):
   # handle multiple exceptions
   # TypeError and ZeroDivisionError
   pass

except:
   # handle all other exceptions
   pass


try:
    num = int(input("Enter a number: "))
    assert num % 2 == 0
except:
    print("Not an even number!")
else:
    reciprocal = 1/num
    print(reciprocal)


# try: 
#     # do main job
# except:
#     # do another - Hereby, you are handling an exception!


class myStrangeError(Exception):
    print("My strange error occurred!")
    
a = "1"
try:
    if a > 5:
        a = a + 100 
    else:
        raise myStrangeError()
except myStrangeError:
    print("My strange error occurred!, so adding only 10")
    a = a + 10
except:
    print("Another error occurred!, so adding 1000")
    a = int(a) + 1000
    



debugging
# https://www.youtube.com/watch?v=w8QHoVam1-I


print('hi')
import pdb; pdb.set_trace()
print('hi2')


# regex

import module

# meta characters
[] . ^ $ * + ? {} () \ |

special sequences
https://pynative.com/python-regex-special-sequences-and-character-classes/

https://regex101.com/

import re

# findall
string = 'hello 12 hi 89. Howdy 34'
pattern = '\d+'
result = re.findall(pattern, string) 
print(result)


# search
string = "My iPython is fun aPython"

# check if 'Python' is at the beginning
match = re.search('Python', string)
match.group()

string = '39801 356, 2102 1111'

# Three digit number followed by space followed by two digit number
pattern = '(\d{3}) (\d{2})'

# match variable contains a Match object.
match = re.search(pattern, string) 

# Output: 801 35

match.group(1)
match.group(2)
match.group(1, 2)
match.groups()

#match
result = re.match("[1-9]{1,3}", "hi 12345 bhey 135456")
result = re.match("[0-9]{1,3}", "12345 dfdf 456")


# split
string = 'Twelve:12 Eighty nine:89.'
pattern = '\d+'
result = re.split(pattern, string) 
print(result)


# sub

# multiline string
string = 'abc 12\
de 23 \n f45 6'

# matches all whitespace characters
pattern = '\s+'
# empty string
replace = ''
new_string = re.sub(pattern, replace, string) 
print(new_string)

flags
ignore_case



text = "Dorj Bat. He was born on 17th of February, 1995\
Has got bachelor degree. Phone: 99779977, uses Mobicom. \
Graduated 3rd secondary school of Khovd aimag. \
Citizen ID KhO95021701. PIN code of Khan bank card is 0011\
Lives with his wife and two children."


citizenId = re.findall("\w{2,3}(?=\d{8})\d{8}",text)
region    = re.findall(r"(Khovd|Uvs|Zavkhan)",text)
mobileOperator = re.findall(r"(Mobicom|Unitel|Skytel|G-mobile)",text)
mobile_number  = re.findall("(?<=\s)\d{8}",text)
bank           = re.findall("\s(\w*)\sbank",text)
nChild         = re.findall("\s(\w*)\schildren",text)



4w
    

Markdown
import pandas as pd
Excel


# get all sheet content 
# choose sheet
# choose cell
# loop through cells
# choose last column and row
# update cell value
# save and save many sheets
# write formula
# hide, freeze, format cells


openpyxl
xlrd
xlwt


https://realpython.com/openpyxl-excel-spreadsheets-python/

https://xlsxwriter.readthedocs.io/
https://www.datacamp.com/community/tutorials/python-excel-tutorial
https://www.dataquest.io/blog/excel-and-pandas/
https://www.python-excel.org/




fileName = "C:/Users/sugarkhuu/Documents/python/repo/Introduction_Python/week2/data.xlsx"

xl = pd.ExcelFile(fileName)
df = xl.parse('Sheet1')

writer = pd.ExcelWriter('example.xlsx', engine='xlsxwriter')
yourData.to_excel(writer, 'Sheet1') # Multiple sheets


with pd.ExcelWriter('multiplesheet.xlsx', engine='xlsxwriter') as writer:
    df.to_excel(writer, sheet_name='Sheet')
    df2.to_excel(writer, sheet_name='Sheet2')
    df3.to_excel(writer, sheet_name='Sheet3')
writer.save()


from openpyxl import load_workbook
wb = load_workbook(fileName)
print(wb.sheetnames)
anotherSheet = wb.active
sheet = wb['Sheet1']
print(sheet['A1'].value)
c = sheet['B3']
print('Row No.:', c.row)
print('Column Letter:', c.column)
print('Coordinates of cell:', c.coordinate)
print('Sheet Title:',sheet.title)
print(sheet.cell(row=1, column=2).value)


from openpyxl.utils import get_column_letter, column_index_from_string

print('Column Letter:', get_column_letter(1))
print('Column Index:', column_index_from_string('A'))


# Print row per row
for cellObj in sheet['A1':'C3']:
      for cell in cellObj:
              print(cell.coordinate, cell.value)


print('Max Rows:', sheet.max_row)
print('Max Columns:', sheet.max_column)
df = pd.DataFrame(sheet.values)


from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import *
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)


import xlrd
workbook = xlrd.open_workbook("C:/Users/sugarkhuu/Documents/python/repo/Introduction_Python/week2/data.xlsx")
worksheet = workbook.sheet_by_name('Sheet1')
worksheet = workbook.sheet_by_index(0)
worksheet.cell(1, 1).value

visibility


import xlwt
book = xlwt.Workbook(encoding="utf-8")
sheet1 = book.add_sheet("Python Sheet 1")
sheet1.write(0, 0, "This is the First Cell of the First Sheet")
book.save("spreadsheet.xls")

# Initialize a workbook
book = xlwt.Workbook()

# Add a sheet to the workbook
sheet1 = book.add_sheet("Sheet1")

# The data
cols = ["A", "B", "C", "D", "E"]
txt = [0,1,2,3,4]

# Loop over the rows and columns and fill in the values
for num in range(5):
      row = sheet1.row(num)
      for index, col in enumerate(cols):
          value = txt[index] + num
          row.write(index, value)

# Save the result
book.save("test.xls")


# pdf

# import file with content
# edit or add pages
# 

pypdf2
tabular-py

python-docx

https://automatetheboringstuff.com/chapter13/
https://theautomatic.net/2019/05/24/3-ways-to-scrape-tables-from-pdfs-with-python/
https://towardsdatascience.com/scraping-table-data-from-pdf-files-using-a-single-line-in-python-8607880c750
https://automatetheboringstuff.com/chapter13/
https://www.datacamp.com/community/tutorials/reading-and-editing-pdfs-and-word-documents-from-python


html later (xml - where is my result?)


latex


5w

Matplotlib, plotly, seaborn, dash (Web?)


# Matplotlib

# Import the necessary packages and modules
import matplotlib.pyplot as plt
import numpy as np

# Prepare the data
x = np.linspace(0, 10, 100)

# Plot the data
plt.plot(x, x, label='linear')

# Add a legend
plt.legend()

# Show the plot
plt.show()


import matplotlib.pyplot as plt
fig = plt.figure()
ax = fig.add_subplot(111)
ax.plot([1, 2, 3, 4], [10, 20, 25, 30], color='lightblue', linewidth=3)
ax.scatter([0.3, 3.8, 1.2, 2.5], [11, 25, 9, 26], color='darkgreen', marker='^')
ax.set_xlim(0.5, 4.5)
plt.show()

import matplotlib.pyplot as plt
plt.plot([1, 2, 3, 4], [10, 20, 25, 30], color='lightblue', linewidth=3)
plt.scatter([0.3, 3.8, 1.2, 2.5], [11, 25, 9, 26], color='darkgreen', marker='^')
plt.xlim(0.5, 4.5)
plt.show()


# Import the necessary packages and modules
import matplotlib.pyplot as plt
import numpy as np

# Create a Figure
fig = plt.figure()

# Set up Axes
ax = fig.add_subplot(111)

# Scatter the data
ax.scatter(np.linspace(0, 1, 5), np.linspace(0, 5, 5))

# Show the plot
plt.show()


# Import `pyplot` from `matplotlib`
import matplotlib.pyplot as plt

# Initialize the plot
fig = plt.figure(figsize=(25,10))
ax1 = fig.add_subplot(121)
ax2 = fig.add_subplot(122)

# or replace the three lines of code above by the following line: 
#fig, (ax1, ax2) = plt.subplots(1,2, figsize=(20,10))

# Plot the data
ax1.bar([1,2,3],[3,4,5])
ax2.barh([0.5,1,2.5],[0,1,2])

# Show the plot
plt.show()



# Import `pyplot` from `matplotlib`
import matplotlib.pyplot as plt

# Initialize the plot
fig = plt.figure()
ax1 = fig.add_subplot(131)
ax2 = fig.add_subplot(132)
ax3 = fig.add_subplot(133)

# Plot the data
ax1.bar([1,2,3],[3,4,5])
ax2.barh([0.5,1,2.5],[0,1,2])
ax2.axhline(0.45)
ax2.fill()
ax1.axvline(0.65)
ax3.scatter(x,y)

# Show the plot
plt.show()


# Import PdfPages
from matplotlib.backends.backend_pdf import PdfPages

# Initialize the pdf file
pp = PdfPages('multipage.pdf')

# Save the figure to the file
pp.savefig()

# Close the file
pp.close()

https://www.machinelearningplus.com/plots/matplotlib-tutorial-complete-guide-python-plot-examples/
https://www.w3schools.com/python/matplotlib_intro.asp
https://matplotlib.org/stable/tutorials/index.html



https://www.activestate.com/blog/plotting-data-in-python-matplotlib-vs-plotly/
https://towardsdatascience.com/matplotlib-vs-plotly-express-which-one-is-the-best-library-for-data-visualization-7a96dbe3ff09
https://towardsdatascience.com/matplotlib-vs-seaborn-vs-plotly-f2b79f5bddb



# Plotly tutorial
https://plotly.com/python/
https://www.kaggle.com/kanncaa1/plotly-tutorial-for-beginners
https://www.geeksforgeeks.org/python-plotly-tutorial/
https://neptune.ai/blog/plotly-python-tutorial-for-machine-learning-specialists


https://pauliacomi.com/2020/06/07/plotly-v-bokeh.html

# scatter, bar, histogram, line, subplots, legend, axis label, title, markers, 


Bokeh
Dash



6w
Web scraping (beautifulsoup, selenium)


7w
Control mouse, screenshot, send email, 

pip install pyautogui
https://pyautogui.readthedocs.io/en/latest/
https://automatetheboringstuff.com/chapter18/


send email
https://stackabuse.com/how-to-send-emails-with-gmail-using-python/


8w

Glimpse into machine learning

Train, test split. Train a models. Rally them. Cross validate. Pick best. 
Simple, Machine learning, AI

Additional
Scipy and numpy
Image
https://www.analyticsvidhya.com/blog/2014/12/image-processing-python-basics/
https://neptune.ai/blog/image-processing-in-python-algorithms-tools-and-methods-you-should-know



Algo trading
https://algotrading101.com/learn/interactive-brokers-python-api-native-guide/
https://www.quora.com/Which-platforms-allow-me-to-plug-in-trading-algorithms-written-in-Python-to-start-trading-and-not-just-backtesting
https://www.datacamp.com/community/tutorials/finance-python-trading
https://towardsdatascience.com/how-to-create-a-fully-automated-ai-based-trading-system-with-python-708503c1a907
https://subscription.packtpub.com/book/big-data-and-business-intelligence/9781784394516/8/ch08lvl1sec70/list-of-trading-platforms-with-public-api

