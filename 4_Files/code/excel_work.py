#%% Working with EXCEL files

# How to install packages?
# pip install package_name

#%% environment
import os
import sys

from settings import *
sys.path.append(r'' + main_dir + '\code')


# set main dir
os.chdir(main_dir+'/'+resource_dir)

#%% Read
import pandas as pd
xlsx = pd.ExcelFile('data.xlsx') # import data file as object
print(xlsx.sheet_names)
df = xlsx.parse("Sheet1") # convert to pandas dataframe
print(df)


# openpyxl
from openpyxl import load_workbook

wb = load_workbook("data.xlsx") # import data file as object 
wb.sheetnames                   # sheets
ws = wb["double"]               # activate sheet "double" as ws
ws.title                        # sheet name
ws["D10"].value                 # cell value by cell name
ws.cell(row=10, column=4).value # cell value by cell location
ws["A1:C2"]
ws["A1:C2"][1][2].value         # cell value from selected range
ws["A:B"][1][1].value           # cell value from selected range

# row and column range
row1  = ws[1:2][1]  # row row 
row12 = ws[1][1:3]    # row col when row is only one
                   
# iter_rows, iter_cols
for value in ws.iter_rows(min_row=1,max_row=4,
                             min_col=1,max_col=5,
                             values_only=True):  # 
    print(value)

for value in ws.iter_cols(min_row=1,max_row=4,
                             min_col=1,max_col=5,
                             values_only=True): # values_only=True
    print(value)

# rows, columns
# loop by rows
for row in ws.rows:
    print(row) 

for row in ws.rows:
    print("{} and {}".format(row[1].value,row[3].value))

# loop by columns
for col in ws.columns:
    print("{} and {}".format(col[1].value,col[3].value))

# loop through cells   
for cell in ws[1]:
    print(cell.value)

# dumping to json - restore/dump concepts in database
import json
info = {}

# Using the values_only because you want to return the cells' values
for row in ws.iter_rows(min_row=2,min_col=1,max_col=4,
                           values_only=True):
    info_id = row[0]
    # dict
    info_i = { 
        "firstName": row[1],
        "lastName": row[2],
        "citizenId": row[3]
    }
    info[info_id] = info_i

with open('../' + result_dir + '/' + 'info.json', 'w') as myfile:
    json.dump(info, myfile, sort_keys=True, indent=10) # 

# find min/max row and cols
ws.min_row
ws.min_column
ws.max_row
ws.max_column


#%% Write
from openpyxl import Workbook
# redefine result_dir for ease
result_dir = '../' + result_dir

wb = Workbook() # instantiate a workbook
ws = wb.active  # select first sheet of wb as sheet

ws["A1"] = "This is"
ws["B1"] = "an apple."

ws["A1"].value
wb.save(result_dir + '/' + "sample.xlsx")         # positional

wb = load_workbook(filename=result_dir + '/' + "sample.xlsx")
ws = wb['Sheet']
ws['A1'].value


cell = ws["A1"]
cell.value = "hey"
wb.save(result_dir + '/' + "sample.xlsx")


# create sheet
wb = load_workbook(result_dir + '/' + "sample.xlsx")
wb.create_sheet("secondSheet",index=None)
wb.save(result_dir + '/' + "sample.xlsx")

# remove sheet
wb = load_workbook(result_dir + '/' + "sample.xlsx")
wb.sheetnames
wb.remove(wb['secondSheet'])
wb.save(result_dir + '/' + "sample.xlsx")

# copy sheet
wb = load_workbook(result_dir + '/' + "sample.xlsx")
wb.sheetnames
wb.copy_worksheet(wb['Sheet'])
wb.sheetnames
ws = wb['Sheet Copy']
ws.title = 'diffSheet'
wb.sheetnames
wb.save(result_dir + '/' + "sample.xlsx")

# formula
wb = load_workbook("data.xlsx")
wb.sheetnames
wb.copy_worksheet(wb['Sheet1'])
ws = wb['Sheet1 Copy']
ws.title = 'formula'
wb.sheetnames
ws["E12"] = '=AVERAGE(E2:E11)'
wb.save(filename="data.xlsx")

# formula flexible
wb = load_workbook(filename="data.xlsx")
wb.sheetnames
wb.copy_worksheet(wb['Sheet1'])
ws = wb['Sheet1 Copy']
ws.title = 'formula2'

last_row = ws.max_row
age_col  = 5

ws.cell(last_row+1,age_col).value = '=AVERAGE(E2:E' + str(last_row) + ')'
wb.save(filename=result_dir + os.sep + "data.xlsx")

# , multiple sheets if pandas
import pandas as pd
xl = pd.ExcelFile("data.xlsx")
df = xl.parse('Sheet1')
with pd.ExcelWriter(result_dir + '/' + 'multiplesheet.xlsx', engine='xlsxwriter') as writer:
    df.to_excel(writer, sheet_name='Sheet1')
    df.to_excel(writer, sheet_name='Sheet2')
    df.to_excel(writer, sheet_name='Sheet3')
writer.save()


# Sources
# https://realpython.com/openpyxl-excel-spreadsheets-python/
# https://xlsxwriter.readthedocs.io/
# https://www.datacamp.com/community/tutorials/python-excel-tutorial
# https://www.dataquest.io/blog/excel-and-pandas/
# https://www.python-excel.org/

